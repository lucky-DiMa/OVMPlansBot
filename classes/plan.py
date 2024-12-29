from copy import copy
from typing import List
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from help import sort_dict
from mongo_connector import mongo_db


class Plan:
    def __init__(self, user_id: int, text: str, date: str):
        self._user_id = user_id
        self._text = text
        self._date = date

    @property
    def text(self):
        return self._text

    @property
    def user_id(self):
        return self._user_id

    @property
    def date(self):
        return self._date

    @text.setter
    def text(self, text: str):
        self._text = text
        mongo_db["Plans"].update_one({"$and": [{"user_id": self.user_id}, {"date": self.date}]},
                                     {"$set": {"text": text}})

    @classmethod
    def get_by_date_and_user_id(cls, user_id: int, date: str):
        dict_plan = mongo_db["Plans"].find_one({"$and": [{"user_id": user_id}, {"date": date}]})
        if dict_plan is None:
            return None
        return cls(dict_plan["user_id"], dict_plan["text"], dict_plan["date"])

    @classmethod
    def get_by_date(cls, date: str):
        list_of_dict_plans = mongo_db["Plans"].find({"date": date})
        list_of_plans = []
        for dict_plan in list_of_dict_plans:
            list_of_plans.append(cls(dict_plan["user_id"], dict_plan["text"], dict_plan["date"]))
        return list_of_plans

    @classmethod
    def get_by_user_id(cls, user_id: int):
        list_of_dict_plans = mongo_db["Plans"].find({"user_id": user_id})
        list_of_plans = []
        for dict_plan in list_of_dict_plans:
            list_of_plans.append(cls(dict_plan["user_id"], dict_plan["text"], dict_plan["date"]))
        return list_of_plans

    @classmethod
    def create(cls, user_id: int, text: str, date: str):
        mongo_db["Plans"].insert_one({"user_id": user_id, "text": text, "date": date})
        return cls.get_by_date_and_user_id(user_id, date)

    @classmethod
    def create_plans_table(cls, str_date: str, sheets: List[str] | None = None):
        if not sheets:
            sheets = mongo_db["Locations"].find_one({})["list"]

        base_font = Font('Calibri', size=11)
        section_font = Font('Calibri', bold=True, size=11)
        no_information_font = Font('Calibri', color='e60000', size=11)
        base_border = Border(Side('thin'), Side('thin'))
        last_border = Border(Side('thin'), Side('thin'), bottom=Side('thin'))
        section_border = Border(Side('thin'), Side('thin'), Side('thin'), Side('thin'))
        alignment = Alignment(horizontal='center')

        from classes import PlansBotUser
        plans_dict = {}
        for location in sheets:
            plans_dict[location] = {"ID": [], "Сотрудник": [], "Отдел": [], "Место": []}
        for user in PlansBotUser.get_all():
            if user.location not in sheets or user.state in ['TYPING FULLNAME', 'CHOOSING LOCATION',
                                                                        'CHOOSING SECTION',
                                                                        'WAITING FOR REG CONFIRMATION']:
                continue
            plans_dict[user.location]["ID"].append(str(user.id))
            plans_dict[user.location]["Сотрудник"].append(user.fullname)
            plans_dict[user.location]["Отдел"].append(user.section)
            plan = Plan.get_by_date_and_user_id(user.id, str_date)

            if plan is None:
                plans_dict[user.location]["Место"].append("Нет информации")
            elif plan.text.startswith('На выезде'):
                plans_dict[user.location]["Место"].append(" ".join(plan.text.split()[2:]))
            else:
                plans_dict[user.location]["Место"].append(plan.text)
        for location in sheets:
            plans_dict[location] = sort_dict(plans_dict[location], 'Сотрудник')
            plans_dict[location] = sort_dict(plans_dict[location], 'Отдел', True)
        wb = Workbook()
        for location in sheets:
            wb.create_sheet(title=location, index=None)
            sheet = wb[location]
            now_section = None
            indent = 1
            for i in range(len(plans_dict[location]["ID"])):
                if i != 0 and now_section != plans_dict[location]['Отдел'][i]:
                    for l in 'ABC':
                        sheet[f'{l}{i + indent}'].border = copy(last_border)
                    sheet.merge_cells(f'A{i + indent}:C{i + indent}')
                    indent += 1
                if i == 0 or now_section != plans_dict[location]['Отдел'][i]:
                    now_section = plans_dict[location]['Отдел'][i]
                    sheet[f'A{i + indent}'].value = now_section
                    sheet[f'A{i + indent}'].font = copy(section_font)
                    sheet[f'A{i + indent}'].border = copy(section_border)
                    sheet.merge_cells(f'A{i + indent}:C{i + indent}')
                    indent += 1
                    sheet[f'A{i + indent}'].value = 'ID'
                    sheet[f'B{i + indent}'].value = 'Сотрудник'
                    sheet[f'C{i + indent}'].value = 'Место'
                    for l in 'ABC':
                        sheet[f'{l}{i + indent}'].font = copy(section_font)
                        sheet[f'{l}{i + indent}'].border = copy(section_border)
                    indent += 1
                sheet[f'A{i + indent}'].value = plans_dict[location]['ID'][i]
                sheet[f'B{i + indent}'].value = plans_dict[location]['Сотрудник'][i]
                sheet[f'C{i + indent}'].value = plans_dict[location]['Место'][i]
                for l in 'ABC':
                    sheet[f'{l}{i + indent}'].font = copy(base_font)
                    if sheet[f'{l}{i + indent}'].value == 'Нет информации':
                        sheet[f'{l}{i + indent}'].font = copy(no_information_font)
                    sheet[f'{l}{i + indent}'].border = copy(base_border)
                    if i == len(plans_dict[location]["ID"]) - 1:
                        sheet[f'{l}{i + indent}'].border = copy(last_border)
            try:
                sheet.column_dimensions['A'].width = max(list(map(len, plans_dict[location]['ID']))) + 2
                sheet.column_dimensions['B'].width = max(list(map(len, plans_dict[location]['Сотрудник']))) + 2
                sheet.column_dimensions['C'].width = max(list(map(len, plans_dict[location]['Место']))) + 2
            except ValueError:
                pass
            for i in range(len(plans_dict[location]['ID']) + indent):
                for l in 'ABC':
                    sheet[f'{l}{i + 1}'].alignment = copy(alignment)
        for sheet in wb:
            wb.remove_sheet(sheet)
            break
        wb.save('./plans.xlsx')

    @classmethod
    def delete_all_by_user_id(cls, user_id: int, sheets: List[str] | None = None):
        mongo_db["Plans"].delete_many({"user_id": user_id})


if __name__ == '___main__0:':
    Plan.create_plans_table('12.10.2023')
